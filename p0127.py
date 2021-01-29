import xlrd
import openpyxl
from datetime import datetime
from DBUtils.PooledDB import PooledDB
import pymysql
import traceback
import time


# 匹配关键词
def pipei(word):
    for e in gjc:
        flag = True
        keys = e.split()
        for k in keys:
            if str(word).find(k) == -1:
                flag = False
                break
        if flag:
            return e
    return None


def get_db():
    # 连接池中获取结果
    connect = pool.connection()

    # 使用cursor()方法获取操作游标
    cursor = connect.cursor()

    # SQL 查询语句
    find_sql = "SELECT keyword,web,country,keym,create_date FROM t0127"

    try:
        # 执行SQL语句
        cursor.execute(find_sql)
        # 获取所有记录列表
        results = cursor.fetchall()
        # print(results)
        return results

    except Exception:
        print("查询数据异常", Exception)
        # 如果发生错误则回滚
        connect.rollback()

    connect.close()
    cursor.close()


# 插数据库
def insert_db(list):
    # 连接池中获取结果
    connect = pool.connection()

    # 使用cursor()方法获取操作游标
    cursor = connect.cursor()

    create_date = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    esql = 'INSERT INTO t0127 (keyword,web,country,keym,create_date) VALUES'
    for e in list:
        esql += " ('%s', '%s', '%s', '%s', '%s')," % (
            e['key'], e['web'], e['country'], e['keym'], create_date)

    esql = esql[:-1] + ';'
    try:
        cursor.execute(esql)
        connect.commit()
    except:
        wrong_sql.append(esql)
        print(esql)
        traceback.print_exc()


# 插excel
# def insert_xlxs(key, web, country, keym, i):
#     if keym:
#         sheet_result.cell(i, 1).value = key
#         sheet_result.cell(i, 2).value = web
#         sheet_result.cell(i, 3).value = country
#         sheet_result.cell(i, 4).value = keym
#     else:
#         sheet_result.cell(i, 1).value = key
#         sheet_result.cell(i, 2).value = web
#         sheet_result.cell(i, 3).value = country


if __name__ == '__main__':
    # init
    time_all = time.time()
    pool = PooledDB(pymysql, 5, host="amourerp.c6ipn3pop4w4.rds.cn-north-1.amazonaws.com.cn", user='oweo56EWO',
                    passwd='fdg2309$!GG', db='test_erp', port=3306, charset="utf8")
    wrong_sql = []
    cost_list = {}
    out_excel = False
    out_db = True

    print(datetime.now())
    gjc = []
    value_list = {}
    insert_list = {}

    # open xlxs
    timer0 = time.process_time()  # cpu执行时间
    timer00 = time.time()  # 程序执行时间：cpu+IO+休眠时间

    workbook = xlrd.open_workbook('wzgjc.xlsx')
    workbook_gjc = xlrd.open_workbook('gjc.xlsx')
    # workbook_result = openpyxl.Workbook()
    # sheet_result = workbook_result.create_sheet(index=0)

    cost_list['open xlxs cpu'] = time.process_time() - timer0
    cost_list['open xlxs'] = time.time() - timer00

    '''
        读gjc.xlsx进入内存gjc
    '''
    table0 = workbook_gjc.sheet_by_index(0)
    rows = table0.nrows
    for row in range(rows):
        value = table0.cell_value(row, 0).lower()
        gjc.append(value)

    '''
        读db原有数据进入内存value_list
    '''
    print('开始读取数据库...')
    timer_read_db = time.time()
    result = get_db()
    for e in result:
        value_list[e[0]] = {'key': e[0], 'web': e[1], 'country': e[2], 'keym': e[3], 'changed': False}
    cost_list['read db'] = time.time() - timer_read_db
    db_nums = len(value_list)
    print('数据库内已有数据：' + str(db_nums) + '条')
    # 读取wzgjc.xlxs数据进入内存value_list（去重）
    timer1 = time.process_time()
    timer11 = time.time()

    sheets = workbook.sheet_names()
    # nsheets = workbook.nsheets
    i = 2
    for sheet in sheets:
        print('读到sheet页:' + sheet)
        country = sheet[-2:]
        web = sheet[:-3]
        table = workbook.sheet_by_name(sheet)

        rows = table.nrows
        cols = table.ncols
        for row in range(0, rows):
            value = str(table.cell_value(row, 0))

            keym = pipei(value)
            if keym:
                keym = pymysql.escape_string(keym)
            value = pymysql.escape_string(value)
            if value in value_list:
                insert_list[value] = {'key': value, 'web': web, 'country': country, 'keym': keym, 'changed': True}
            else:
                insert_list[value] = {'key': value, 'web': web, 'country': country, 'keym': keym, 'changed': False}

            i += 1
    print('wzgjc.xlxs数据：' + str(i - 2) + '条')
    print('开始插入数据库...')
    # if i > 100000:
    #     break

    cost_list['read xlxs dedup cpu'] = time.process_time() - timer1
    cost_list['read xlxs dedup'] = time.time() - timer11

    '''
        insert_list数据写入数据库或excel
    '''
    timer2 = time.process_time()
    timer22 = time.time()
    i = 2
    list = []
    nsql = 10000  # 一条sql同时写入的数据数量
    for e in insert_list:
        if insert_list[e]['changed']:
            continue
        list.append(insert_list[e])
        if i % nsql == 0:
            # 每nsql条数据插入一次数据库
            insert_db(list)
            list = []

        # insert_xlxs(value_list[e]['key'], value_list[e]['web'], value_list[e]['country'], value_list[e]['keym'], i)
        i += 1
    if len(list) is not 0:
        insert_db(list)
    print('新增有效数据：' + str(i - 2) + '条')

    cost_list['insert db cpu'] = time.process_time() - timer2
    cost_list['insert db'] = time.time() - timer22

    timer3 = time.process_time()
    timer33 = time.time()
    # workbook_result.save('result_gjc2.xlsx')
    cost_list['save xlxs cpu'] = time.process_time() - timer3
    cost_list['save xlxs'] = time.time() - timer33
    cost_list['all'] = time.time() - time_all

    print('耗时：' + str(cost_list))
    print(datetime.now())
